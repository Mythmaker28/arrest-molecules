#!/usr/bin/env python3
"""
Script pour générer API_Calculations_Full.xlsx
Workbook multi-feuilles avec calculs API détaillés

Author: Tommy Lepesteur
License: MIT
Version: 1.0
Date: October 2025

Requirements: pip install openpyxl pandas numpy
Usage: python generate_excel_workbook.py
"""

import pandas as pd
import numpy as np
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("⚠️  openpyxl non installé. Installation...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

def create_formulas_sheet(wb):
    """Sheet 1: Définition des formules et variables"""
    ws = wb.create_sheet("1_Formulas", 0)
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # En-tête
    ws['A1'] = "Arrest Potency Index (API) - Définition Mathématique"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Formule principale
    ws['A3'] = "Formule API (absolu)"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    
    ws['A4'] = "API = [(1/K_d) × τ_residence] / [t_onset × EC₅₀]"
    ws['A4'].font = Font(bold=True, size=11)
    
    # Variables
    ws['A6'] = "Variable"
    ws['B6'] = "Description"
    ws['C6'] = "Unités"
    ws['D6'] = "Plage typique"
    for cell in ['A6', 'B6', 'C6', 'D6']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
    
    variables = [
        ["K_d", "Constante de dissociation à l'équilibre", "nM", "0.001 – 100,000"],
        ["τ_residence", "Temps de résidence sur la cible (1/k_off)", "min", "0.1 – 10,000"],
        ["t_onset", "Temps d'apparition de l'effet (50% E_max)", "min", "0.1 – 10,000"],
        ["EC₅₀", "Concentration efficace médiane", "nM", "0.001 – 100,000"],
        ["API_abs", "Arrest Potency Index absolu", "nM⁻²", "10⁻⁶ – 10⁴"],
        ["API_rel", "API relatif (normalisé à salvinorin A)", "—", "10⁻⁵ – 10"]
    ]
    
    for i, var in enumerate(variables, start=7):
        ws[f'A{i}'] = var[0]
        ws[f'B{i}'] = var[1]
        ws[f'C{i}'] = var[2]
        ws[f'D{i}'] = var[3]
    
    # Normalisation
    ws['A15'] = "Normalisation"
    ws['A15'].font = header_font
    ws['A15'].fill = header_fill
    
    ws['A16'] = "API_relatif = API_absolu / API_salvinorin_A"
    ws['A17'] = "Référence : Salvinorin A API_absolu = 6.95 nM⁻²"
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20

def create_parameters_sheet(wb):
    """Sheet 2: Paramètres extraits de Compound_Properties_Database.csv"""
    # Lire le CSV
    df = pd.read_csv("Compound_Properties_Database.csv")
    
    ws = wb.create_sheet("2_Parameters")
    
    # Sélectionner colonnes pertinentes pour API
    cols_api = ['Compound_Name', 'K_d_nM', 'tau_residence_min', 't_onset_min', 
                'EC50_nM', 'API_absolute', 'API_relative']
    df_api = df[cols_api].copy()
    
    # Renommer pour clarté
    df_api.columns = ['Compound', 'K_d (nM)', 'τ_residence (min)', 't_onset (min)', 
                      'EC₅₀ (nM)', 'API_absolu (nM⁻²)', 'API_relatif']
    
    # Écrire dans Excel
    for r in dataframe_to_rows(df_api, index=False, header=True):
        ws.append(r)
    
    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15

def create_examples_sheet(wb):
    """Sheet 3: Exemples de calculs détaillés"""
    ws = wb.create_sheet("3_Step_by_Step_Examples")
    
    # En-tête
    ws['A1'] = "Exemples de Calculs API Détaillés"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Exemple Salvinorin A
    ws['A3'] = "Exemple 1: Salvinorin A (composé de référence)"
    ws['A3'].font = Font(bold=True, size=12, color="0070C0")
    
    steps_salv = [
        ["Étape", "Calcul", "Valeur"],
        ["1. Paramètres d'entrée", "", ""],
        ["", "K_d", "1.8 nM"],
        ["", "τ_residence", "25 min"],
        ["", "t_onset", "1 min"],
        ["", "EC₅₀", "2 nM"],
        ["2. Calcul numérateur", "(1/K_d) × τ_residence", "= (1/1.8) × 25 = 13.89"],
        ["3. Calcul dénominateur", "t_onset × EC₅₀", "= 1 × 2 = 2.0"],
        ["4. API absolu", "Numérateur / Dénominateur", "= 13.89 / 2.0 = 6.95 nM⁻²"],
        ["5. API relatif", "API_abs / API_salv_A", "= 6.95 / 6.95 = 1.00"]
    ]
    
    row = 4
    for step in steps_salv:
        ws[f'A{row}'] = step[0]
        ws[f'B{row}'] = step[1]
        ws[f'C{row}'] = step[2]
        if "Paramètres" in step[0] or "Calcul" in step[0] or "API" in step[0]:
            ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Exemple Rapamycin
    row += 2
    ws[f'A{row}'] = "Exemple 2: Rapamycin (onset lent)"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="0070C0")
    
    steps_rapa = [
        ["Étape", "Calcul", "Valeur"],
        ["1. Paramètres d'entrée", "", ""],
        ["", "K_d", "0.1 nM"],
        ["", "τ_residence", "120 min"],
        ["", "t_onset", "1440 min (24h)"],
        ["", "EC₅₀", "1 nM"],
        ["2. Calcul numérateur", "(1/K_d) × τ_residence", "= (1/0.1) × 120 = 1200"],
        ["3. Calcul dénominateur", "t_onset × EC₅₀", "= 1440 × 1 = 1440"],
        ["4. API absolu", "Numérateur / Dénominateur", "= 1200 / 1440 = 0.833 nM⁻²"],
        ["5. API relatif", "API_abs / API_salv_A", "= 0.833 / 6.95 = 0.12"]
    ]
    
    row += 1
    for step in steps_rapa:
        ws[f'A{row}'] = step[0]
        ws[f'B{row}'] = step[1]
        ws[f'C{row}'] = step[2]
        if "Paramètres" in step[0] or "Calcul" in step[0] or "API" in step[0]:
            ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30

def create_monte_carlo_sheet(wb):
    """Sheet 4: Résultats Monte Carlo simulation"""
    ws = wb.create_sheet("4_Monte_Carlo_Results")
    
    # En-tête
    ws['A1'] = "Monte Carlo Uncertainty Quantification (10,000 iterations)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    # Headers
    headers = ['Compound', 'API_median', 'API_mean', 'API_std', 'CI_2.5%', 'CI_97.5%', 'Confidence']
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Données simulées (cohérentes avec code Python)
    np.random.seed(42)
    compounds_data = [
        ('Salvinorin A', 1.8, 25, 1, 2, 'VERY HIGH'),
        ('Paclitaxel', 0.9, 720, 360, 10, 'HIGH'),
        ('Rapamycin', 0.1, 120, 1440, 1, 'MODERATE'),
        ('Capsaicin', 700, 45, 5, 700, 'MODERATE-HIGH'),
        ('Tetrodotoxin', 10, 360, 3, 10, 'HIGH'),
        ('Resveratrol', 50000, 1, 720, 50000, 'LOW')
    ]
    
    reference_api = 6.95
    
    row = 4
    for compound, kd, tau, onset, ec50, conf in compounds_data:
        # Simulation Monte Carlo simplifiée
        kd_samples = np.random.lognormal(np.log(kd), 0.15, 10000)
        tau_samples = np.random.lognormal(np.log(tau), 0.15, 10000)
        onset_samples = np.random.lognormal(np.log(onset), 0.15, 10000)
        ec50_samples = np.random.lognormal(np.log(ec50), 0.15, 10000)
        
        api_samples = ((1/kd_samples) * tau_samples) / (onset_samples * ec50_samples)
        api_rel_samples = api_samples / reference_api
        
        ws[f'A{row}'] = compound
        ws[f'B{row}'] = round(np.median(api_rel_samples), 5)
        ws[f'C{row}'] = round(np.mean(api_rel_samples), 5)
        ws[f'D{row}'] = round(np.std(api_rel_samples), 5)
        ws[f'E{row}'] = round(np.percentile(api_rel_samples, 2.5), 5)
        ws[f'F{row}'] = round(np.percentile(api_rel_samples, 97.5), 5)
        ws[f'G{row}'] = conf
        row += 1
    
    # Ajuster largeurs
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18

def create_provenance_sheet(wb):
    """Sheet 5: Métadonnées et provenance"""
    ws = wb.create_sheet("5_Provenance")
    
    ws['A1'] = "Métadonnées et Provenance"
    ws['A1'].font = Font(bold=True, size=14)
    
    metadata = [
        ["Champ", "Valeur"],
        ["Titre", "API Calculations Full - Molecular Arrest Framework"],
        ["Auteur", "Tommy Lepesteur"],
        ["Email", "tommy.lepesteur@hotmail.fr"],
        ["ORCID", "0009-0009-0577-9563"],
        ["Version", "1.0"],
        ["Date de création", datetime.now().strftime("%Y-%m-%d")],
        ["Licence (données)", "CC-BY 4.0"],
        ["Licence (code)", "MIT"],
        ["Source", "Compound_Properties_Database.csv + Python_Code_API_Monte_Carlo.py"],
        ["Manuscrit", "Molecular Arrest in Biological Regulation (Lepesteur 2025)"],
        ["GitHub", "https://github.com/Mythmaker28/arrest-molecules"],
        ["Seed Monte Carlo", "42 (reproductibilité)"],
        ["Itérations MC", "10,000 per compound"],
        ["Logiciel", "Python 3.8+ (numpy, pandas, openpyxl)"],
        ["Notes", "Fichier généré automatiquement par generate_excel_workbook.py"]
    ]
    
    row = 3
    for field, value in metadata:
        ws[f'A{row}'] = field
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

def main():
    """Génération du workbook Excel complet"""
    print("Génération de API_Calculations_Full.xlsx...")
    
    wb = Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut
    
    create_formulas_sheet(wb)
    print("  ✓ Sheet 1: Formulas")
    
    create_parameters_sheet(wb)
    print("  ✓ Sheet 2: Parameters")
    
    create_examples_sheet(wb)
    print("  ✓ Sheet 3: Step-by-Step Examples")
    
    create_monte_carlo_sheet(wb)
    print("  ✓ Sheet 4: Monte Carlo Results")
    
    create_provenance_sheet(wb)
    print("  ✓ Sheet 5: Provenance")
    
    # Sauvegarder
    output_path = "API_Calculations_Full.xlsx"
    wb.save(output_path)
    print(f"\n✅ Fichier sauvegardé : {output_path}")
    print(f"   5 feuilles créées avec succès")
    print(f"   Taille : ~{round(len(open(output_path, 'rb').read())/1024, 1)} KB")

if __name__ == "__main__":
    main()

